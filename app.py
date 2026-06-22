import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import re


# ────────────────────────────────────────────────
#  설정
# ────────────────────────────────────────────────
st.set_page_config(
    page_title="수열 학습 RPG",
    page_icon="🎮",
    layout="wide"
)

TOTAL_LESSONS = 19
NO_MENTOR = "멘토 없음 / 혼자 풀었음"

ALL_STUDENTS = {
    "2학년 5반": ["곽민석","김다율","김제니아","백종우","이시원",
                  "조건희","최민준","최은호","김서진","김채은",
                  "김하윤","남수정","박찬비","양은서","이수연",
                  "전유영","정민경","조은서","홍기담"],
    "2학년 6반": ["김채민","양수현","양준환","여민겸","윤영민",
                  "이도윤","전재훈","전진우","김예원","김유주",
                  "류효연","송민경","안혜준","안효은","이다인",
                  "이연서","이예은","전하은","정연아"]
}

LEVEL_TABLE = [
    (0,  2,  1, "🌱", "수학 새싹",     "#FECACA", "#DC2626"),
    (3,  5,  2, "📚", "성실한 학습자", "#FED7AA", "#EA580C"),
    (6,  9,  3, "⚔️", "베테랑 수학자", "#FEF08A", "#CA8A04"),
    (10, 14, 4, "🌟", "수열 마스터",   "#BBF7D0", "#16A34A"),
    (15, 19, 5, "👑", "전설의 수학왕", "#BFDBFE", "#2563EB"),
]

def get_level_info(count):
    for mn, mx, lv, icon, title, bg, fg in LEVEL_TABLE:
        if mn <= count <= mx:
            return lv, icon, title, bg, fg, count - mn, mx - mn + 1
    return 5, "👑", "전설의 수학왕", "#BFDBFE", "#2563EB", TOTAL_LESSONS, TOTAL_LESSONS

# ────────────────────────────────────────────────
#  CSS — 흰 배경, 큰 글씨, 단순 카드
# ────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;600;700;900&display=swap');

html, body, [class*="css"] {
    font-family: 'Noto Sans KR', sans-serif;
    color: #111827;
}

/* 흰 배경 */
.stApp { background: #FFFFFF; }
[data-testid="stSidebar"] {
    background: #F9FAFB;
    border-right: 1px solid #E5E7EB;
}

/* 탭 */
[data-testid="stTabs"] button {
    font-size: 1rem !important;
    font-weight: 600 !important;
    color: #6B7280 !important;
    padding: 0.6rem 1rem !important;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #111827 !important;
    border-bottom: 2px solid #111827 !important;
}

/* 상단 요약 카드 */
.sum-card {
    background: #F9FAFB;
    border: 1px solid #E5E7EB;
    border-radius: 12px;
    padding: 1.2rem 1rem;
    text-align: center;
}
.sum-card .s-val {
    font-size: 2.2rem;
    font-weight: 900;
    color: #111827;
    line-height: 1;
}
.sum-card .s-lbl {
    font-size: 0.9rem;
    color: #6B7280;
    margin-top: 0.4rem;
    font-weight: 500;
}

/* 플레이어 카드 */
.player-card {
    background: #FFFFFF;
    border: 1.5px solid #E5E7EB;
    border-radius: 14px;
    padding: 1.1rem 1.2rem;
    margin-bottom: 0.75rem;
}
.player-card .top-row {
    display: flex;
    align-items: center;
    gap: 0.6rem;
    margin-bottom: 0.5rem;
}
.player-card .p-rank {
    font-size: 1rem;
    color: #9CA3AF;
    min-width: 1.8rem;
    font-weight: 600;
}
.player-card .p-icon { font-size: 1.3rem; }
.player-card .p-name {
    font-size: 1.05rem;
    font-weight: 700;
    color: #111827;
    flex: 1;
}
.lv-badge {
    font-size: 0.78rem;
    font-weight: 700;
    padding: 0.2rem 0.65rem;
    border-radius: 20px;
}
.player-card .p-title {
    font-size: 0.85rem;
    color: #6B7280;
    margin-bottom: 0.55rem;
    padding-left: 2.7rem;
    font-weight: 500;
}
.xp-track {
    background: #F3F4F6;
    border-radius: 6px;
    height: 10px;
    overflow: hidden;
    margin-bottom: 0.35rem;
}
.xp-fill {
    height: 100%;
    border-radius: 6px;
}
.xp-row {
    display: flex;
    justify-content: space-between;
    font-size: 0.8rem;
    color: #6B7280;
    font-weight: 500;
}

/* 멘토 행 */
.mentor-row {
    display: flex;
    align-items: center;
    padding: 0.85rem 0;
    border-bottom: 1px solid #F3F4F6;
    gap: 0.8rem;
}
.mentor-row .mr-rank {
    font-size: 1.1rem;
    min-width: 2.2rem;
    text-align: center;
}
.mentor-row .mr-name {
    font-size: 1rem;
    font-weight: 700;
    color: #111827;
    flex: 1;
}
.mentor-row .mr-bar-wrap {
    flex: 2;
    background: #F3F4F6;
    border-radius: 4px;
    height: 8px;
    overflow: hidden;
}
.mentor-row .mr-bar-fill {
    height: 8px;
    border-radius: 4px;
    background: #2563EB;
}
.mentor-row .mr-score {
    font-size: 0.95rem;
    font-weight: 700;
    color: #2563EB;
    min-width: 3rem;
    text-align: right;
}

/* 레벨 분포 배지 */
.lv-dist-card {
    background: #F9FAFB;
    border: 1.5px solid #E5E7EB;
    border-radius: 12px;
    padding: 1rem;
    text-align: center;
}
.lv-dist-card .ld-icon { font-size: 1.8rem; }
.lv-dist-card .ld-cnt {
    font-size: 2rem;
    font-weight: 900;
    color: #111827;
    margin: 0.2rem 0;
}
.lv-dist-card .ld-lbl {
    font-size: 0.82rem;
    color: #6B7280;
    font-weight: 500;
}

/* 반 카드 */
.class-card {
    background: #F9FAFB;
    border: 1.5px solid #E5E7EB;
    border-radius: 14px;
    padding: 1.4rem 1.5rem;
    margin-bottom: 1rem;
}
.class-card h3 {
    font-size: 1.15rem;
    font-weight: 700;
    color: #111827;
    margin: 0 0 1rem;
}
.class-stat-val {
    font-size: 1.9rem;
    font-weight: 900;
    color: #111827;
}
.class-stat-lbl {
    font-size: 0.82rem;
    color: #6B7280;
    font-weight: 500;
    margin-top: 0.2rem;
}
.progress-track {
    background: #E5E7EB;
    border-radius: 8px;
    height: 14px;
    overflow: hidden;
    margin: 0.8rem 0;
}
.progress-fill {
    height: 14px;
    border-radius: 8px;
    background: #111827;
}

h2, h3 { color: #111827 !important; }
</style>
""", unsafe_allow_html=True)

# ────────────────────────────────────────────────
#  구글 시트 연결
# ────────────────────────────────────────────────
@st.cache_resource
def get_gc():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_dict(
        st.secrets["gcp_service_account"], scope
    )
    return gspread.authorize(creds)

@st.cache_data(ttl=300)
def load_all_responses():
    gc = get_gc()
    try:
        sh = gc.open("수열_폼_응답_통합")
    except Exception:
        return pd.DataFrame()
    all_dfs = []
    for ws in sh.worksheets():
        try:
            raw = ws.get_all_values()
            if len(raw) < 2:
                continue
            headers = raw[0]
            seen = {}
            clean = []
            for h in headers:
                seen[h] = seen.get(h, 0) + 1
                clean.append(h if seen[h] == 1 else f"{h}_{seen[h]}")
            df = pd.DataFrame(raw[1:], columns=clean)
            df = df[df.iloc[:, 0] != ""].copy()
            if df.empty:
                continue
            tab = ws.title
            df["차시"] = tab.split("_", 1)[1] if "_" in tab and not tab.startswith("Form") else tab
            for base in ["본인 이름", "도움 준 멘토 이름"]:
                dup = base + "_2"
                if base in df.columns and dup in df.columns:
                    df[base] = df[base].replace("", pd.NA).fillna(df[dup])
                    df.drop(columns=[dup], inplace=True)
                elif dup in df.columns:
                    df.rename(columns={dup: base}, inplace=True)
            ts = clean[0]
            df.rename(columns={ts: "타임스탬프"}, inplace=True)
            df["타임스탬프"] = pd.to_datetime(df["타임스탬프"], errors="coerce")
            all_dfs.append(df)
        except Exception:
            continue
    return pd.concat(all_dfs, ignore_index=True) if all_dfs else pd.DataFrame()

def medal(i):
    return ["🥇", "🥈", "🥉"][i] if i < 3 else f"{i+1}위"

# ────────────────────────────────────────────────
#  헤더
# ────────────────────────────────────────────────
st.markdown("## 🎮 수열 학습 RPG 대시보드")
st.markdown(
    '<p style="color:#6B7280;font-size:1rem;margin-top:-0.5rem">'
    '대동세무고등학교 · 대수 | 제출할수록 레벨이 올라요!</p>',
    unsafe_allow_html=True
)
st.divider()

# ────────────────────────────────────────────────
#  사이드바
# ────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🔍 필터")
    class_filter = st.multiselect(
        "반 선택",
        options=list(ALL_STUDENTS.keys()),
        default=list(ALL_STUDENTS.keys())
    )
    if st.button("🔄 새로고침"):
        st.cache_data.clear()
        st.rerun()

    st.divider()
    st.markdown("### 🏆 레벨 안내")
    for mn, mx, lv, icon, title, bg, fg in LEVEL_TABLE:
        st.markdown(
            f'<div style="padding:0.4rem 0;font-size:0.9rem">'
            f'{icon} <b>Lv.{lv}</b> {title}<br>'
            f'<span style="color:#9CA3AF;font-size:0.8rem">{mn}~{mx}차시 제출</span></div>',
            unsafe_allow_html=True
        )

# ────────────────────────────────────────────────
#  데이터 로드
# ────────────────────────────────────────────────
df = load_all_responses()

if df.empty:
    st.warning("⚠️ 아직 응답 데이터가 없거나 시트 연결을 확인해주세요.")
    st.stop()

df_f = df[df["반"].isin(class_filter)].copy() if "반" in df.columns and class_filter else df.copy()
def lesson_sort_key(s):
    # 차시명에서 첫 번째 숫자를 추출해서 정렬 기준으로 사용
    # 예: "1차시 - 수열의 뜻" → 1, "11차시 - 합의 기호" → 11
    m = re.search(r'\d+', str(s))
    return int(m.group()) if m else 9999

all_lessons = sorted(
    df_f["차시"].unique().tolist(),
    key=lesson_sort_key
) if "차시" in df_f.columns else []
mentor_col = next((c for c in ["도움 준 멘토 이름", "멘토 이름"] if c in df_f.columns), None)
total_students = sum(len(v) for k, v in ALL_STUDENTS.items() if k in class_filter)
submitted_names = df_f["본인 이름"].nunique() if "본인 이름" in df_f.columns else 0
mentor_count = int(df_f[mentor_col].notna().sum()) if mentor_col else 0

# ────────────────────────────────────────────────
#  상단 요약 지표
# ────────────────────────────────────────────────
c1, c2, c3, c4 = st.columns(4)
for col, val, lbl in [
    (c1, len(df_f), "총 제출 수"),
    (c2, f"{submitted_names}/{total_students}명", "참여 학생"),
    (c3, f"{len(all_lessons)}/{TOTAL_LESSONS}차시", "진행 차시"),
    (c4, mentor_count, "멘토 지목 수"),
]:
    col.markdown(f"""
    <div class="sum-card">
        <div class="s-val">{val}</div>
        <div class="s-lbl">{lbl}</div>
    </div>
    """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ────────────────────────────────────────────────
#  탭
# ────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "📊 전체 현황",
    "⚔️ 개인 레벨",
    "👑 멘토 랭킹",
    "🏰 반별 현황",
])

# ══════════════════════════════════════
#  탭 1: 전체 현황
# ══════════════════════════════════════
with tab1:
    st.markdown("### 차시별 제출 현황")
    if "차시" in df_f.columns:
        lesson_counts = (
            df_f.groupby("차시").size()
            .reset_index(name="제출 수")
            .sort_values("차시")
        )
        fig = px.bar(
            lesson_counts, x="차시", y="제출 수",
            text="제출 수",
            color_discrete_sequence=["#111827"],
        )
        fig.update_traces(textposition="outside", marker_line_width=0)
        fig.update_layout(
            paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
            font=dict(color="#111827", size=13),
            xaxis=dict(tickangle=-45, gridcolor="#F3F4F6", tickfont=dict(size=12)),
            yaxis=dict(gridcolor="#F3F4F6", tickfont=dict(size=12)),
            height=360, margin=dict(t=30, b=60),
        )
        st.plotly_chart(fig, use_container_width=True)

    st.divider()
    st.markdown("### 레벨 분포 현황")
    if "본인 이름" in df_f.columns:
        student_submit_counts = df_f.groupby("본인 이름")["차시"].nunique()
        lv_dist = {lv: 0 for _, _, lv, _, _, _, _ in LEVEL_TABLE}
        for cnt in student_submit_counts.values:
            lv, *_ = get_level_info(cnt)
            lv_dist[lv] = lv_dist.get(lv, 0) + 1

        cols = st.columns(5)
        for i, (mn, mx, lv, icon, title, bg, fg) in enumerate(LEVEL_TABLE):
            cnt = lv_dist.get(lv, 0)
            cols[i].markdown(f"""
            <div class="lv-dist-card">
                <div class="ld-icon">{icon}</div>
                <div class="ld-cnt">{cnt}명</div>
                <div class="ld-lbl">Lv.{lv} {title}</div>
            </div>
            """, unsafe_allow_html=True)

# ══════════════════════════════════════
#  탭 2: 개인 레벨
# ══════════════════════════════════════
with tab2:
    if "본인 이름" not in df_f.columns:
        st.info("본인 이름 데이터를 찾을 수 없습니다.")
    else:
        for class_name in ALL_STUDENTS:
            if class_name not in class_filter:
                continue

            st.markdown(f"### {class_name}")
            students = ALL_STUDENTS[class_name]
            class_df = df_f[df_f["반"] == class_name] if "반" in df_f.columns else df_f

            # 학생별 데이터 수집
            student_data = []
            for name in students:
                s_df = class_df[class_df["본인 이름"] == name]
                cnt = s_df["차시"].nunique()
                submitted_lessons = set(s_df["차시"].unique())
                lv, icon, title, bg, fg, xp_in, xp_need = get_level_info(cnt)
                student_data.append((name, cnt, lv, icon, title, bg, fg, xp_in, xp_need, submitted_lessons))
            student_data.sort(key=lambda x: (-x[2], -x[1]))

            # ── 레벨 카드 ──
            col_a, col_b = st.columns(2)
            for i, (name, cnt, lv, icon, title, bg, fg, xp_in, xp_need, _) in enumerate(student_data):
                pct = min(int(xp_in / xp_need * 100), 100)
                is_max = (lv == 5 and cnt >= 15)
                border = f"border-color:{fg}" if is_max else ""
                suffix = "✨ 전설 달성!" if is_max else f"다음 레벨까지 {xp_need - xp_in} XP"
                card = f"""
                <div class="player-card" style="{border}">
                    <div class="top-row">
                        <span class="p-rank">{medal(i)}</span>
                        <span class="p-icon">{icon}</span>
                        <span class="p-name">{name}</span>
                        <span class="lv-badge" style="background:{bg};color:{fg}">Lv.{lv}</span>
                    </div>
                    <div class="p-title">{title} · {cnt}/{TOTAL_LESSONS} 차시 완료</div>
                    <div class="xp-track">
                        <div class="xp-fill" style="width:{pct}%;background:{fg}"></div>
                    </div>
                    <div class="xp-row">
                        <span>{xp_in}/{xp_need} XP</span>
                        <span>{suffix}</span>
                    </div>
                </div>
                """
                (col_a if i % 2 == 0 else col_b).markdown(card, unsafe_allow_html=True)

            st.markdown("<br>", unsafe_allow_html=True)

            # ── 차시별 제출 현황 표 ──
            st.markdown(f"#### 📋 {class_name} 차시별 제출 현황")

            # 피벗 데이터 생성
            rows = []
            for name, cnt, lv, icon, title, bg, fg, xp_in, xp_need, submitted_lessons in student_data:
                row = {
                    "이름": name,
                    "레벨": f"Lv.{lv} {icon}",
                    "제출 수": f"{cnt}/{len(all_lessons)}",
                }
                for lesson in all_lessons:
                    # 차시명 짧게 (앞 숫자만 or 앞 6자)
                    short = lesson[:6] if len(lesson) > 6 else lesson
                    row[short] = "✅" if lesson in submitted_lessons else "❌"
                rows.append(row)

            pivot_df = pd.DataFrame(rows)
            st.dataframe(pivot_df, hide_index=True, use_container_width=True)

            # ── 미제출 요약 ──
            with st.expander("⚠️ 미제출 상세 보기"):
                any_missing = False
                for name, cnt, lv, icon, title, bg, fg, xp_in, xp_need, submitted_lessons in student_data:
                    missing = [l for l in all_lessons if l not in submitted_lessons]
                    if missing:
                        any_missing = True
                        short_missing = [l[:8] for l in missing]
                        st.markdown(
                            f'<p style="font-size:0.95rem;margin:0.3rem 0">'
                            f'<b>{name}</b> · 미제출 {len(missing)}개: '
                            f'<span style="color:#DC2626">{", ".join(short_missing)}</span></p>',
                            unsafe_allow_html=True
                        )
                if not any_missing:
                    st.markdown('<p style="color:#16A34A;font-weight:700">✅ 전원 모든 차시 제출 완료!</p>', unsafe_allow_html=True)

            # ── 엑셀 다운로드 ──
            import io
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            def make_excel(student_data, all_lessons, class_name):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = class_name[:30]

                # 헤더 스타일
                header_fill = PatternFill("solid", fgColor="111827")
                header_font = Font(color="FFFFFF", bold=True, size=11)
                center = Alignment(horizontal="center", vertical="center")
                thin = Side(style="thin", color="E5E7EB")
                border = Border(left=thin, right=thin, top=thin, bottom=thin)

                # 헤더 행
                headers = ["이름", "레벨", "제출 수"] + [l[:8] for l in all_lessons]
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=h)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = border

                # 데이터 행
                green_fill = PatternFill("solid", fgColor="D1FAE5")
                red_fill   = PatternFill("solid", fgColor="FEE2E2")

                for row_idx, (name, cnt, lv, icon, title, bg, fg, xp_in, xp_need, submitted_lessons) in enumerate(student_data, 2):
                    row_vals = [name, f"Lv.{lv} {icon}", f"{cnt}/{len(all_lessons)}"]
                    for lesson in all_lessons:
                        row_vals.append("✅" if lesson in submitted_lessons else "❌")

                    for col_idx, val in enumerate(row_vals, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=val)
                        cell.alignment = center
                        cell.border = border
                        # 제출 여부 셀 색칠
                        if col_idx > 3:
                            cell.fill = green_fill if val == "✅" else red_fill

                # 열 너비 조정
                ws.column_dimensions["A"].width = 12
                ws.column_dimensions["B"].width = 14
                ws.column_dimensions["C"].width = 8
                for col_idx in range(4, len(headers) + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 9

                # 행 높이
                for row_idx in range(1, len(student_data) + 2):
                    ws.row_dimensions[row_idx].height = 22

                buf = io.BytesIO()
                wb.save(buf)
                buf.seek(0)
                return buf.getvalue()

            excel_data = make_excel(student_data, all_lessons, class_name)
            st.download_button(
                label=f"📥 {class_name} 제출 현황 엑셀 다운로드",
                data=excel_data,
                file_name=f"수열_제출현황_{class_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{class_name}"
            )

            st.markdown("<br>", unsafe_allow_html=True)
# ══════════════════════════════════════
#  탭 3: 멘토 랭킹
# ══════════════════════════════════════
with tab3:
    st.markdown("### 👑 멘토 포인트 랭킹")
    st.markdown(
        '<p style="color:#6B7280;font-size:0.95rem">친구를 도와줄수록 멘토 포인트가 쌓여요!</p>',
        unsafe_allow_html=True
    )

    if mentor_col and "본인 이름" in df_f.columns:
        mentor_df = df_f[
            (df_f[mentor_col] != NO_MENTOR) &
            (df_f[mentor_col] != "") &
            (df_f[mentor_col].notna()) &
            (df_f[mentor_col] != df_f["본인 이름"])
        ]
        mentor_counts = (
            mentor_df.groupby(mentor_col).size()
            .reset_index(name="포인트")
            .sort_values("포인트", ascending=False)
        )

        if mentor_counts.empty:
            st.info("아직 멘토 지목 데이터가 없습니다.")
        else:
            col_list, col_chart = st.columns([2, 3])

            with col_list:
                max_pt = mentor_counts["포인트"].max()
                for i, (_, row) in enumerate(mentor_counts.head(10).iterrows()):
                    name = row[mentor_col]
                    pt = row["포인트"]
                    bar_pct = int(pt / max_pt * 100)
                    st.markdown(f"""
                    <div class="mentor-row">
                        <div class="mr-rank">{medal(i)}</div>
                        <div class="mr-name">{name}</div>
                        <div class="mr-bar-wrap">
                            <div class="mr-bar-fill" style="width:{bar_pct}%"></div>
                        </div>
                        <div class="mr-score">{pt}pt</div>
                    </div>
                    """, unsafe_allow_html=True)

            with col_chart:
                top10 = mentor_counts.head(10)
                fig2 = go.Figure(go.Bar(
                    x=top10["포인트"],
                    y=top10[mentor_col],
                    orientation="h",
                    marker_color="#111827",
                    text=top10["포인트"].astype(str) + "pt",
                    textposition="outside",
                    textfont=dict(color="#111827", size=13),
                ))
                fig2.update_layout(
                    paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
                    font=dict(color="#111827", size=13),
                    xaxis=dict(gridcolor="#F3F4F6"),
                    yaxis=dict(
                        categoryorder="total ascending",
                        tickfont=dict(size=13, color="#111827"),
                    ),
                    height=380, margin=dict(t=10, b=10, l=10, r=60),
                )
                st.plotly_chart(fig2, use_container_width=True)
    else:
        st.info("멘토 이름 컬럼을 찾을 수 없습니다.")

# ══════════════════════════════════════
#  탭 4: 반별 현황
# ══════════════════════════════════════
with tab4:
    st.markdown("### 🏰 반별 퀘스트 현황")

    for class_name in ALL_STUDENTS:
        if class_name not in class_filter:
            continue

        students = ALL_STUDENTS[class_name]
        class_df = df_f[df_f["반"] == class_name] if "반" in df_f.columns else df_f
        total = len(students)

        submitted_set = set(class_df["본인 이름"].unique()) if "본인 이름" in class_df.columns else set()
        submitted_cnt = len([s for s in students if s in submitted_set])
        rate = submitted_cnt / total * 100 if total else 0

        levels = []
        for name in students:
            cnt = class_df[class_df["본인 이름"] == name]["차시"].nunique() if "본인 이름" in class_df.columns else 0
            lv, *_ = get_level_info(cnt)
            levels.append(lv)
        avg_lv = sum(levels) / len(levels) if levels else 0

        st.markdown(f"""
        <div class="class-card">
            <h3>🏰 {class_name}</h3>
            <div style="display:flex;gap:2rem;margin-bottom:1rem">
                <div>
                    <div class="class-stat-val">{submitted_cnt}/{total}명</div>
                    <div class="class-stat-lbl">참여 학생</div>
                </div>
                <div>
                    <div class="class-stat-val">Lv.{avg_lv:.1f}</div>
                    <div class="class-stat-lbl">평균 레벨</div>
                </div>
                <div>
                    <div class="class-stat-val">{rate:.0f}%</div>
                    <div class="class-stat-lbl">참여율</div>
                </div>
            </div>
            <div class="progress-track">
                <div class="progress-fill" style="width:{rate}%"></div>
            </div>
        </div>
        """, unsafe_allow_html=True)

        # 레벨 분포 + 미제출
        col_pie, col_missing = st.columns([2, 3])

        with col_pie:
            lv_dist = {lv: 0 for _, _, lv, _, _, _, _ in LEVEL_TABLE}
            for name in students:
                cnt = class_df[class_df["본인 이름"] == name]["차시"].nunique() if "본인 이름" in class_df.columns else 0
                lv, *_ = get_level_info(cnt)
                lv_dist[lv] = lv_dist.get(lv, 0) + 1

            fig3 = go.Figure(go.Pie(
                labels=[f"Lv.{lv}" for _, _, lv, _, _, _, _ in LEVEL_TABLE],
                values=list(lv_dist.values()),
                hole=0.55,
                marker=dict(colors=["#FECACA","#FED7AA","#FEF08A","#BBF7D0","#BFDBFE"]),
                textfont=dict(color="#111827", size=13),
            ))
            fig3.update_layout(
                paper_bgcolor="#FFFFFF",
                showlegend=True,
                legend=dict(font=dict(color="#111827", size=12)),
                height=260,
                margin=dict(t=10, b=10, l=10, r=10),
                annotations=[dict(
                    text=f"평균<br>Lv.{avg_lv:.1f}",
                    x=0.5, y=0.5, showarrow=False,
                    font=dict(size=15, color="#111827")
                )],
            )
            st.plotly_chart(fig3, use_container_width=True)

        with col_missing:
            not_submitted = [s for s in students if s not in submitted_set]
            if not_submitted:
                st.markdown(
                    f'<p style="font-size:0.95rem;font-weight:700;color:#111827;margin-bottom:0.5rem">'
                    f'⚠️ 미참여 학생 {len(not_submitted)}명</p>',
                    unsafe_allow_html=True
                )
                for name in not_submitted:
                    st.markdown(
                        f'<span style="display:inline-block;background:#F3F4F6;'
                        f'border-radius:20px;padding:0.25rem 0.75rem;margin:0.2rem;'
                        f'font-size:0.9rem;color:#374151">{name}</span>',
                        unsafe_allow_html=True
                    )
            else:
                st.markdown(
                    '<p style="font-size:1rem;font-weight:700;color:#16A34A">✅ 전원 퀘스트 참여 완료!</p>',
                    unsafe_allow_html=True
                )

        st.divider()
